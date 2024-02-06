import openpyxl
from django.shortcuts import render, redirect
from django.views import View
from .forms import ExcelUploadForm, CategoryForm
from django.db import models
from .models import Product, ProductAttributes, Category
import datetime
import csv
from django.contrib import messages
from django.db.utils import IntegrityError
from django.views.generic import ListView, DetailView
from django.shortcuts import get_object_or_404

def home(request):
    return render(request, 'excel_processor/index.html')

class ExcelUploadView(View):
    def get(self, request):
        form = ExcelUploadForm()
        category_form = CategoryForm(request.POST)

        return render(request, 'excel_processor/upload.html', {'excel_form': form, 'category_form': category_form})

    def post(self, request):
        form = ExcelUploadForm(request.POST, request.FILES)
        category_form = CategoryForm(request.POST)

        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active

                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

                for row in sheet.iter_rows(min_row=2):
                    product_name = row[0].value
                    product, created = Product.objects.get_or_create(name=product_name)
                    category_id = request.POST.get('category')
                    if category_id:
                        category = Category.objects.get(pk=category_id)
                        product.category = category
                        product.save()

                    for header, cell in zip(headers[1:], row[1:]):
                        attribute_value = cell.value
                        ProductAttributes.objects.create(product=product, key=header, value=attribute_value)

                return render(request, 'excel_processor/upload_success.html')

            except Exception as e:
                return render(request, 'excel_processor/index.html')
        elif category_form.is_valid():
            try:
                category_form.save()
                messages.success(request, 'Category created successfully.')
                return redirect('upload')
            except IntegrityError:
                messages.error(request, 'A category with this name already exists.')
                return redirect('upload?show_category_form=true')
        else:
            return render(request, 'excel_processor/upload.html', {'excel_form': form, 'category_form': category_form})


def upload_success(request):
    return render(request, 'excel_processor/upload_success.html')

class CategoriesListView(ListView):
    model = Category
    template_name = 'excel_processor/categories.html'
    context_object_name = 'categories'

class ProductsByCategoryListView(ListView):
    model = Product
    template_name = 'excel_processor/products_by_category.html'
    context_object_name = 'products'

    def get_queryset(self):
        self.category = get_object_or_404(Category, pk=self.kwargs['category_id'])
        return Product.objects.filter(category=self.category)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category'] = self.category
        context['category_name'] = self.category.name
        return context

class ProductDetailView(DetailView):
    model = Product
    template_name = 'excel_processor/product_detail.html'
    context_object_name = 'product'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.get_object()
        attributes = ProductAttributes.objects.filter(product=product)
        context['attributes'] = attributes
        return context
